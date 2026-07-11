import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.api.deps import require_role
from app.db.session import get_db
from app.models.pinjaman import Simpanan, Pinjaman
from app.models.transaksi import Order
from app.models.user import User, UserRole, AnggotaProfile

router = APIRouter(prefix="/laporan", tags=["laporan"])


def _header(ws, kolom: list[str]):
    ws.append(kolom)
    for cell in ws[1]:
        cell.font = Font(bold=True)


@router.get("/export")
def export_laporan(
    dari_tanggal: date | None = Query(default=None),
    sampai_tanggal: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.PENGURUS)),
):
    dari = datetime.combine(dari_tanggal, datetime.min.time()) if dari_tanggal else None
    sampai = datetime.combine(sampai_tanggal, datetime.max.time()) if sampai_tanggal else None

    simpanan_q = db.query(Simpanan)
    pinjaman_q = db.query(Pinjaman)
    order_q = db.query(Order).filter(Order.is_synthetic == False)  # noqa: E712
    if dari:
        simpanan_q = simpanan_q.filter(Simpanan.created_at >= dari)
        pinjaman_q = pinjaman_q.filter(Pinjaman.created_at >= dari)
        order_q = order_q.filter(Order.created_at >= dari)
    if sampai:
        simpanan_q = simpanan_q.filter(Simpanan.created_at <= sampai)
        pinjaman_q = pinjaman_q.filter(Pinjaman.created_at <= sampai)
        order_q = order_q.filter(Order.created_at <= sampai)

    simpanan_list = simpanan_q.order_by(Simpanan.created_at).all()
    pinjaman_list = pinjaman_q.order_by(Pinjaman.created_at).all()
    order_list = order_q.order_by(Order.created_at).all()

    anggota_map = {p.id: p for p in db.query(AnggotaProfile).all()}
    user_map = {u.id: u for u in db.query(User).all()}

    def nama_anggota(anggota_id):
        profile = anggota_map.get(anggota_id)
        if not profile:
            return "-"
        user = user_map.get(profile.user_id)
        return user.full_name if user else "-"

    wb = Workbook()

    ws_ringkasan = wb.active
    ws_ringkasan.title = "Ringkasan"
    total_setor = sum(float(s.nominal) for s in simpanan_list if s.jenis.value == "setor")
    total_tarik = sum(float(s.nominal) for s in simpanan_list if s.jenis.value == "tarik")
    total_pinjaman_disalurkan = sum(
        float(p.nominal) for p in pinjaman_list if p.status.value in ("approved", "lunas")
    )
    total_transaksi_marketplace = sum(float(o.total) for o in order_list)
    ws_ringkasan.append(["Laporan Keuangan REKA"])
    ws_ringkasan.append([f"Periode: {dari_tanggal or 'semua'} s.d. {sampai_tanggal or 'semua'}"])
    ws_ringkasan.append([])
    _header(ws_ringkasan, ["Metrik", "Nilai (Rp)"])
    ws_ringkasan.append(["Total Setor Simpanan", total_setor])
    ws_ringkasan.append(["Total Tarik Simpanan", total_tarik])
    ws_ringkasan.append(["Total Pinjaman Disalurkan", total_pinjaman_disalurkan])
    ws_ringkasan.append(["Total Transaksi Marketplace", total_transaksi_marketplace])

    ws_simpanan = wb.create_sheet("Simpanan")
    _header(ws_simpanan, ["Tanggal", "Anggota", "Jenis", "Nominal"])
    for s in simpanan_list:
        ws_simpanan.append([s.created_at.strftime("%Y-%m-%d %H:%M"), nama_anggota(s.anggota_id), s.jenis.value, float(s.nominal)])

    ws_pinjaman = wb.create_sheet("Pinjaman")
    _header(ws_pinjaman, ["Tanggal", "Anggota", "Nominal", "Tujuan", "Status", "Risk Level", "Skor Gabungan"])
    for p in pinjaman_list:
        ws_pinjaman.append([
            p.created_at.strftime("%Y-%m-%d %H:%M"), nama_anggota(p.anggota_id), float(p.nominal),
            p.tujuan, p.status.value, p.risk_level or "-", float(p.skor_gabungan) if p.skor_gabungan else "-",
        ])

    ws_order = wb.create_sheet("Transaksi Marketplace")
    _header(ws_order, ["Tanggal", "Status", "Total", "Kurir"])
    for o in order_list:
        ws_order.append([o.created_at.strftime("%Y-%m-%d %H:%M"), o.status.value, float(o.total), o.kurir or "-"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"laporan-reka-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
